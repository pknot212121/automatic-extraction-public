import fitz
import os
import pytesseract
from PIL import Image, ImageEnhance
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import re
import cv2
import numpy as np
import time
import win32com.client
from usuwanie_pieczątek import *

def get_images_from_pdf(pdf,folder="C:/Users/knotp/OneDrive/Pulpit/automatic_extraction",dpi=300,brightness_factor=1.5):
    doc = fitz.open(folder+'/'+pdf)
    names=[]
    for i,page in enumerate(doc[2:-2]):
        zoom=dpi/72
        mat=fitz.Matrix(zoom,zoom)
        pix=page.get_pixmap(matrix=mat,alpha=False)
        img=Image.frombytes("RGB",[pix.width,pix.height],pix.samples)
        np_img = usuń_pieczątkę(np.array(img))
        resized=cv2.resize(np_img,None,fx=1,fy=1)
        print(pytesseract.image_to_string(np_img,lang='pol'))
        image_name=f'image_{i}.jpg'
        cv2.imwrite(image_name,resized)
        names.append(image_name)

def extract_text(pdf,folder,dpi=300,brightness_factor=1.2):
    doc = fitz.open(folder+'/'+pdf)
    i=1
    all_text=""
    for page in doc[:-1]:
        zoom=dpi/72
        mat=fitz.Matrix(zoom,zoom)
        pix=page.get_pixmap(matrix=mat,alpha=False)
        img=Image.frombytes("RGB",[pix.width,pix.height],pix.samples)
        enchancer=ImageEnhance.Brightness(img)
        img_enchanced=enchancer.enhance(brightness_factor)
        np_img = usuń_pieczątkę(np.array(img_enchanced))
        cv2.imwrite('temp.jpg',np_img)
        start_time=time.time()
        text=pytesseract.image_to_string(np_img,lang='pol')
        end_time=time.time()
        print(text)
        print(f"Czas odczytu: {end_time-start_time}")
        all_text+=text
        i+=1
    doc.close()
    return all_text

def create_excel_file(filename):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sample Sheet"
    wb.save(filename)

def input_data_into_excel(name,lines):
    detect_and_close_excel(name)
    try:
        workbook = openpyxl.load_workbook(name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i,line in enumerate(lines):
        for j,cena in enumerate(line):
            sheet.cell(row=i+1,column=j+1,value=cena)
    workbook.save(name)
    open_excel_visibly(name)

def detect_and_close_excel(file_path):
    if not os.path.exists(file_path):
        print(f"Plik {file_path} nie istnieje.")
        return
    try:
        excel = win32com.client.GetObject(Class="Excel.Application")
    except:
        print("Excel nie jest uruchomiony.")
        return
    for workbook in excel.Workbooks:
        if os.path.abspath(workbook.FullName).lower() == os.path.abspath(file_path).lower():
            workbook.Close(SaveChanges=False)
            return
    print(f"Plik {file_path} nie jest obecnie otwarty w Excelu.")

def open_excel_visibly(file_path):
    if not os.path.exists(file_path):
        print(f"Plik {file_path} nie istnieje.")
        return
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True
        excel.DisplayScrollBars = True
        excel.Workbooks.Open(os.path.abspath(file_path))
    except Exception as e:
        print(f"Wystąpił błąd podczas otwierania pliku: {str(e)}")

def write_string_to_file(file_path, content):
    with open(file_path, 'w',encoding='utf-8') as file:
        file.write(content)

def extract_text_from_folder(folder):
    full = ""
    page_counter = 0
    sum=0
    for root, dirs, files in os.walk(folder):
        sum+=len(files)
    licznik=0
    for root, dirs, files in os.walk(folder):
        for i,filename in enumerate(files):
            if filename.endswith(".pdf") or filename.endswith(".PDF"):
                licznik+=1
                file_path = os.path.join(root, filename)
                relative_path = os.path.relpath(file_path, folder)
                print(f"-- {licznik} na {sum} plików --")
                print(f"\n\nPAGE {page_counter}\n")
                print(f"\n\nFILE: {relative_path}\n")
                full += f"\n\nPAGE {page_counter}\n"
                full += f"\n\nFILE: {relative_path}\n"
                t = extract_text(filename, root)
                print(t)
                full+=t
                page_counter += 1

    return full
def extract_text_from_szczegoly_folder():
    folder="D:\\pobrane_szczegoly"
    pliki= [os.path.join(folder,f) for f in os.listdir(folder) if os.path.isfile(os.path.join(folder,f))]
    pliki.sort(key=lambda x: os.path.getmtime(x))
    fulltext=""
    for i,plik in enumerate(pliki):
        doc = fitz.open(plik)
        strona = doc.load_page(0)
        tekst = strona.get_text()
        fulltext+=tekst
        fulltext+="\n----||----\n"
    return fulltext

def read_from_text(text_file):
    with open(text_file, 'r',encoding='utf-8') as file:
        text = file.read()
    return text

# get_images_from_pdf("akn_79-2023_472-2023.pdf")